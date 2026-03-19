"""
Unit tests for core.reporting.excel_export.

Uses synthetic data so no real strategy files are needed.
"""
from __future__ import annotations

import io
import zipfile
from dataclasses import dataclass, field
from datetime import date, timedelta

import numpy as np
import pandas as pd
import pytest

from core.reporting.excel_export import (
    correlations_export_filename,
    export_correlations,
    export_loo_result,
    export_mc_result,
    export_portfolio,
    export_raw_data,
    loo_export_filename,
    mc_export_filename,
    portfolio_export_filename,
    raw_data_export_filename,
)


# ── Fixtures ──────────────────────────────────────────────────────────────────

def _make_dates(n: int):
    start = pd.Timestamp("2022-01-03")
    return pd.date_range(start, periods=n, freq="B")


def _make_strategy(name="ES_Trend", status="Live", contracts=1, symbol="ES", sector="Index"):
    from core.data_types import Strategy
    return Strategy(
        name=name, folder=__import__("pathlib").Path("/fake"),
        status=status, contracts=contracts, symbol=symbol, sector=sector,
    )


def _make_portfolio(n_strats=3, n_days=252):
    from core.data_types import PortfolioData
    from core.config import AppConfig

    dates  = _make_dates(n_days)
    names  = [f"Strat_{i}" for i in range(n_strats)]
    pnl    = pd.DataFrame(np.random.randn(n_days, n_strats) * 500, index=dates, columns=names)
    closed = pd.DataFrame(np.random.randn(n_days, n_strats) * 300, index=dates, columns=names)

    strategies = [_make_strategy(n) for n in names]

    # Minimal summary_metrics DataFrame
    summary = pd.DataFrame(
        {
            "expected_annual_profit": [10_000.0] * n_strats,
            "profit_since_oos_start": [5_000.0]  * n_strats,
            "max_drawdown_isoos":     [0.15]      * n_strats,
            "sharpe_isoos":           [0.8]       * n_strats,
        },
        index=names,
    )

    portfolio = PortfolioData(
        strategies=strategies,
        daily_pnl=pnl,
        closed_trades=closed,
        summary_metrics=summary,
    )
    config = AppConfig()
    return portfolio, config


def _make_imported(n_strats=2, n_days=100):
    from core.data_types import ImportedData, Strategy
    from pathlib import Path

    dates = _make_dates(n_days)
    names = [f"S_{i}" for i in range(n_strats)]
    rng   = np.random.default_rng(0)

    def _df():
        return pd.DataFrame(rng.standard_normal((n_days, n_strats)) * 100, index=dates, columns=names)

    trades = pd.DataFrame({
        "strategy": ["S_0", "S_1"] * 5,
        "date":     dates[:10],
        "position": ["Long", "Short"] * 5,
        "pnl":      rng.standard_normal(10) * 200,
        "mae":      rng.standard_normal(10) * 50,
        "mfe":      rng.standard_normal(10) * 50,
    })

    return ImportedData(
        daily_m2m=_df(), closed_trade_pnl=_df(),
        in_market_long=_df(), in_market_short=_df(),
        trades=trades,
        strategies=[Strategy(n, Path("/fake"), "") for n in names],
    )


def _make_mc_result(with_scenarios=True):
    from core.data_types import MCResult
    scenarios = pd.DataFrame({
        "profit": np.random.randn(100) * 5000 + 20000,
        "max_drawdown_pct": np.random.rand(100) * 0.3,
    }) if with_scenarios else None
    return MCResult(
        starting_equity=100_000.0,
        expected_profit=20_000.0,
        risk_of_ruin=0.05,
        max_drawdown_pct=0.20,
        sharpe_ratio=1.2,
        return_to_drawdown=2.5,
        scenarios_df=scenarios,
    )


def _is_valid_xlsx(data: bytes) -> bool:
    """Check the bytes are a valid xlsx (it's a ZIP with xl/ content)."""
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            return any(n.startswith("xl/") for n in zf.namelist())
    except Exception:
        return False


def _sheet_names(data: bytes) -> list[str]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data))
    return wb.sheetnames


# ── export_portfolio ──────────────────────────────────────────────────────────

class TestExportPortfolio:
    def test_returns_bytes(self):
        portfolio, config = _make_portfolio()
        result = export_portfolio(portfolio, config)
        assert isinstance(result, bytes)
        assert len(result) > 1000

    def test_valid_xlsx(self):
        portfolio, config = _make_portfolio()
        assert _is_valid_xlsx(export_portfolio(portfolio, config))

    def test_sheet_names(self):
        portfolio, config = _make_portfolio()
        sheets = _sheet_names(export_portfolio(portfolio, config))
        assert "Summary" in sheets
        assert "Portfolio Equity" in sheets
        assert "Export Info" in sheets

    def test_summary_has_strategy_rows(self):
        from openpyxl import load_workbook
        portfolio, config = _make_portfolio(n_strats=4)
        wb = load_workbook(io.BytesIO(export_portfolio(portfolio, config)))
        ws = wb["Summary"]
        # Row 1 = title, row 2 = blank, row 3 = header, rows 4+ = data
        data_rows = [row for row in ws.iter_rows(min_row=4, values_only=True) if any(c is not None for c in row)]
        assert len(data_rows) == 4

    def test_empty_summary_handled(self):
        from core.data_types import PortfolioData
        from core.config import AppConfig
        portfolio = PortfolioData(
            strategies=[],
            daily_pnl=pd.DataFrame(),
            closed_trades=pd.DataFrame(),
            summary_metrics=pd.DataFrame(),
        )
        result = export_portfolio(portfolio, AppConfig())
        assert _is_valid_xlsx(result)


# ── export_raw_data ───────────────────────────────────────────────────────────

class TestExportRawData:
    def test_returns_bytes(self):
        imp = _make_imported()
        assert isinstance(export_raw_data(imp), bytes)

    def test_valid_xlsx(self):
        assert _is_valid_xlsx(export_raw_data(_make_imported()))

    def test_sheet_names(self):
        sheets = _sheet_names(export_raw_data(_make_imported()))
        assert "Daily M2M" in sheets
        assert "Closed Trades" in sheets
        assert "In Market Long" in sheets
        assert "In Market Short" in sheets
        assert "Trades" in sheets
        assert "Export Info" in sheets

    def test_daily_m2m_row_count(self):
        from openpyxl import load_workbook
        n_days = 50
        imp = _make_imported(n_days=n_days)
        wb = load_workbook(io.BytesIO(export_raw_data(imp)))
        ws = wb["Daily M2M"]
        # title row + blank + header + data rows
        data_rows = [r for r in ws.iter_rows(min_row=4, values_only=True) if any(c is not None for c in r)]
        assert len(data_rows) == n_days

    def test_empty_trades_handled(self):
        imp = _make_imported()
        imp.trades = pd.DataFrame()
        assert _is_valid_xlsx(export_raw_data(imp))


# ── export_mc_result ──────────────────────────────────────────────────────────

class TestExportMCResult:
    def test_returns_bytes(self):
        assert isinstance(export_mc_result(_make_mc_result()), bytes)

    def test_valid_xlsx(self):
        assert _is_valid_xlsx(export_mc_result(_make_mc_result()))

    def test_sheets_with_scenarios(self):
        sheets = _sheet_names(export_mc_result(_make_mc_result(with_scenarios=True)))
        assert "MC Summary" in sheets
        assert "Scenario Distribution" in sheets

    def test_sheets_without_scenarios(self):
        sheets = _sheet_names(export_mc_result(_make_mc_result(with_scenarios=False)))
        assert "MC Summary" in sheets
        assert "Scenario Distribution" not in sheets


# ── export_loo_result ─────────────────────────────────────────────────────────

class TestExportLOOResult:
    def _make_loo(self):
        return pd.DataFrame({
            "strategy":       ["Strat_A", "Strat_B", "Strat_C"],
            "delta_profit":   [-1000.0, 500.0, -200.0],
            "delta_sharpe":   [-0.1, 0.05, -0.02],
            "delta_drawdown": [0.01, -0.02, 0.0],
            "delta_rtd":      [-0.2, 0.1, -0.05],
            "delta_ror":      [0.01, -0.01, 0.0],
        })

    def test_returns_bytes(self):
        assert isinstance(export_loo_result(self._make_loo()), bytes)

    def test_valid_xlsx(self):
        assert _is_valid_xlsx(export_loo_result(self._make_loo()))


# ── export_correlations ───────────────────────────────────────────────────────

class TestExportCorrelations:
    def _make_corr(self, n=5):
        rng = np.random.default_rng(0)
        names = [f"S{i}" for i in range(n)]
        data = rng.random((n, n))
        data = (data + data.T) / 2
        np.fill_diagonal(data, 1.0)
        return pd.DataFrame(data, index=names, columns=names)

    def test_returns_bytes(self):
        assert isinstance(export_correlations(self._make_corr()), bytes)

    def test_valid_xlsx(self):
        assert _is_valid_xlsx(export_correlations(self._make_corr()))

    def test_mode_in_sheet_name(self):
        sheets = _sheet_names(export_correlations(self._make_corr(), mode="Drawdown"))
        assert any("Drawdown" in s for s in sheets)

    def test_diagonal_is_one(self):
        from openpyxl import load_workbook
        corr = self._make_corr(3)
        wb = load_workbook(io.BytesIO(export_correlations(corr)))
        ws = list(wb.worksheets)[0]
        # B2, C3, D4 should be 1.0 (diagonal)
        assert abs(float(ws.cell(2, 2).value) - 1.0) < 0.01


# ── Filename helpers ──────────────────────────────────────────────────────────

class TestFilenameHelpers:
    def test_portfolio(self):
        name = portfolio_export_filename()
        assert name.endswith(".xlsx") and "portfolio" in name

    def test_raw_data(self):
        name = raw_data_export_filename()
        assert name.endswith(".xlsx") and "raw" in name

    def test_mc(self):
        assert mc_export_filename().endswith(".xlsx")

    def test_loo(self):
        assert loo_export_filename().endswith(".xlsx")

    def test_correlations(self):
        assert correlations_export_filename("Normal").endswith(".xlsx")
