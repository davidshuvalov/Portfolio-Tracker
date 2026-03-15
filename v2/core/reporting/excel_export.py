"""
Excel export — generates .xlsx reports from in-memory portfolio data.

Functions:
  export_portfolio(portfolio_data, config)   -> bytes   Computed-output report
  export_raw_data(imported_data)             -> bytes   Raw imported DataFrames
  export_mc_result(mc_result, label)         -> bytes   Monte Carlo results
  export_loo_result(loo_result, base_*args)  -> bytes   Leave-One-Out table
  export_correlations(corr_df, mode)         -> bytes   Correlation matrix

Requires openpyxl (listed in pyproject.toml dependencies).
"""

from __future__ import annotations

import io
from datetime import date
from typing import TYPE_CHECKING

import pandas as pd

if TYPE_CHECKING:
    from core.config import AppConfig
    from core.data_types import ImportedData, MCResult, PortfolioData


# ── Style helpers ─────────────────────────────────────────────────────────────

def _make_workbook():
    try:
        from openpyxl import Workbook
        return Workbook()
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel export. "
            "Install it with: pip install openpyxl"
        )


def _header_style(ws, row: int = 1) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment
    fill = PatternFill("solid", fgColor="1565C0")
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[row]:
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _autofit(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 45)


def _df_to_sheet(ws, df: pd.DataFrame, title: str | None = None) -> None:
    from openpyxl.styles import Font
    row_offset = 0
    if title:
        ws.append([title])
        ws.cell(row=1, column=1).font = Font(bold=True, size=13)
        ws.append([])
        row_offset = 2

    ws.append(list(df.columns))
    _header_style(ws, row=row_offset + 1)

    for _, row in df.iterrows():
        ws.append([v.item() if hasattr(v, "item") else v for v in row])

    _autofit(ws)


def _save(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Computed-output portfolio export ─────────────────────────────────────────

def export_portfolio(
    portfolio_data: "PortfolioData",
    config: "AppConfig",
) -> bytes:
    """
    Export computed portfolio analytics to .xlsx.

    Sheets:
      Summary         — summary_metrics (80+ columns) for every strategy
      Portfolio Equity — cumulative portfolio P&L curve
      Export Info     — metadata
    """
    wb = _make_workbook()
    wb.remove(wb.active)

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws_summary = wb.create_sheet("Summary")
    summary = portfolio_data.summary_metrics

    if not summary.empty:
        # Prepend Status + Contracts from Strategy objects
        status_map    = {s.name: s.status    for s in portfolio_data.strategies}
        contracts_map = {s.name: s.contracts for s in portfolio_data.strategies}
        df = summary.copy().reset_index().rename(columns={"index": "Strategy"})
        df.insert(1, "Status",    [status_map.get(n, "")    for n in df["Strategy"]])
        df.insert(2, "Contracts", [contracts_map.get(n, 1) for n in df["Strategy"]])

        # Convert date columns to strings for Excel compatibility
        for col in df.columns:
            if pd.api.types.is_datetime64_any_dtype(df[col]):
                df[col] = df[col].dt.strftime("%Y-%m-%d")
            elif df[col].dtype == object:
                # Try to detect date objects
                try:
                    sample = df[col].dropna().iloc[0]
                    if hasattr(sample, "strftime"):
                        df[col] = df[col].apply(lambda v: v.strftime("%Y-%m-%d") if v is not None and hasattr(v, "strftime") else v)
                except (IndexError, TypeError):
                    pass

        _df_to_sheet(ws_summary, df, title="Strategy Summary")
    else:
        ws_summary.append(["No strategy metrics available."])

    # ── Portfolio equity curve ────────────────────────────────────────────────
    ws_equity = wb.create_sheet("Portfolio Equity")
    if not portfolio_data.daily_pnl.empty:
        port_pnl = portfolio_data.daily_pnl.sum(axis=1)
        port_equity = port_pnl.cumsum()
        eq_df = pd.DataFrame({
            "Date":       port_equity.index.strftime("%Y-%m-%d"),
            "Daily PnL ($)":  port_pnl.values.round(2),
            "Equity ($)": port_equity.values.round(2),
        })
        _df_to_sheet(ws_equity, eq_df, title="Portfolio Equity Curve")
    else:
        ws_equity.append(["No portfolio data available."])

    # ── Metadata ─────────────────────────────────────────────────────────────
    ws_meta = wb.create_sheet("Export Info")
    ws_meta.append(["Portfolio Tracker v2 — Computed Output Export"])
    ws_meta.append(["Generated",         str(date.today())])
    ws_meta.append(["Live strategies",   len(portfolio_data.strategies)])
    ws_meta.append(["Period (years)",    config.portfolio.period_years])
    _autofit(ws_meta)

    return _save(wb)


# ── Raw data export ───────────────────────────────────────────────────────────

def export_raw_data(imported_data: "ImportedData") -> bytes:
    """
    Export the raw imported DataFrames to .xlsx.

    Sheets:
      Daily M2M     — daily mark-to-market PnL (dates × strategies)
      Closed Trades — daily closed-trade PnL   (dates × strategies)
      In Market Long
      In Market Short
      Trades        — individual trade records
    """
    wb = _make_workbook()
    wb.remove(wb.active)

    def _matrix_to_sheet(ws, df: pd.DataFrame, title: str) -> None:
        out = df.copy()
        out.index = out.index.strftime("%Y-%m-%d")
        out = out.reset_index().rename(columns={"index": "Date"})
        out = out.round(2)
        _df_to_sheet(ws, out, title=title)

    ws1 = wb.create_sheet("Daily M2M")
    _matrix_to_sheet(ws1, imported_data.daily_m2m, "Daily Mark-to-Market PnL")

    ws2 = wb.create_sheet("Closed Trades")
    _matrix_to_sheet(ws2, imported_data.closed_trade_pnl, "Daily Closed-Trade PnL")

    ws3 = wb.create_sheet("In Market Long")
    _matrix_to_sheet(ws3, imported_data.in_market_long, "In-Market Long PnL")

    ws4 = wb.create_sheet("In Market Short")
    _matrix_to_sheet(ws4, imported_data.in_market_short, "In-Market Short PnL")

    ws5 = wb.create_sheet("Trades")
    if not imported_data.trades.empty:
        trades = imported_data.trades.copy()
        if pd.api.types.is_datetime64_any_dtype(trades.get("date", pd.Series(dtype="object"))):
            trades["date"] = trades["date"].dt.strftime("%Y-%m-%d")
        trades = trades.round(2)
        _df_to_sheet(ws5, trades, title="Individual Trade Records")
    else:
        ws5.append(["No trade-level data available."])

    ws_meta = wb.create_sheet("Export Info")
    start, end = imported_data.date_range
    ws_meta.append(["Portfolio Tracker v2 — Raw Data Export"])
    ws_meta.append(["Generated",    str(date.today())])
    ws_meta.append(["Strategies",   len(imported_data.strategy_names)])
    ws_meta.append(["Date Range",   f"{start} → {end}"])
    ws_meta.append(["Trading Days", len(imported_data.daily_m2m)])
    ws_meta.append(["Trades",       len(imported_data.trades)])
    _autofit(ws_meta)

    return _save(wb)


# ── Monte Carlo export ────────────────────────────────────────────────────────

def export_mc_result(mc_result: "MCResult", label: str = "Portfolio") -> bytes:
    """Export MC summary + scenario distribution to .xlsx."""
    wb = _make_workbook()
    wb.remove(wb.active)

    ws_summary = wb.create_sheet("MC Summary")
    ws_summary.append(["Portfolio Tracker v2 — Monte Carlo Results"])
    ws_summary.append(["Target",              label])
    ws_summary.append(["Generated",           str(date.today())])
    ws_summary.append([])
    ws_summary.append(["Metric", "Value"])
    from openpyxl.styles import Font
    ws_summary.cell(row=5, column=1).font = Font(bold=True)
    ws_summary.cell(row=5, column=2).font = Font(bold=True)

    rows = [
        ("Starting Equity ($)",    round(mc_result.starting_equity, 2)),
        ("Expected Annual Profit ($)", round(mc_result.expected_profit, 2)),
        ("Risk of Ruin",           f"{mc_result.risk_of_ruin:.3%}"),
        ("Max Drawdown (median)",  f"{mc_result.max_drawdown_pct:.3%}"),
        ("Sharpe Ratio",           round(mc_result.sharpe_ratio, 3)),
        ("Return / Drawdown",      round(mc_result.return_to_drawdown, 3)),
    ]
    for r in rows:
        ws_summary.append(list(r))
    _autofit(ws_summary)

    if mc_result.scenarios_df is not None and not mc_result.scenarios_df.empty:
        ws_scenarios = wb.create_sheet("Scenario Distribution")
        df = mc_result.scenarios_df.copy().round(4)
        # Format pct columns
        for col in df.columns:
            if "pct" in col.lower():
                df[col] = df[col].apply(lambda v: f"{v:.2%}")
        _df_to_sheet(ws_scenarios, df, title="Monte Carlo Scenarios")

    return _save(wb)


# ── Leave-One-Out export ──────────────────────────────────────────────────────

def export_loo_result(
    loo_result: "pd.DataFrame",
    base_profit: float = 0.0,
    base_sharpe: float = 0.0,
) -> bytes:
    """Export LOO analysis table to .xlsx."""
    wb = _make_workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Leave-One-Out")
    ws.append(["Portfolio Tracker v2 — Leave-One-Out Analysis"])
    ws.append(["Generated",         str(date.today())])
    ws.append(["Base Exp. Profit",  round(base_profit, 2)])
    ws.append(["Base Sharpe",       round(base_sharpe, 3)])
    ws.append([])

    df = loo_result.copy()
    for col in df.select_dtypes(include="number").columns:
        df[col] = df[col].round(4)

    _df_to_sheet(ws, df, title=None)
    _autofit(ws)

    return _save(wb)


# ── Correlations export ───────────────────────────────────────────────────────

def export_correlations(corr_df: "pd.DataFrame", mode: str = "Normal") -> bytes:
    """Export a correlation matrix DataFrame to .xlsx."""
    wb = _make_workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet(f"Correlations ({mode})")
    from openpyxl.styles import Font, PatternFill, Alignment

    header_fill = PatternFill("solid", fgColor="1565C0")
    header_font = Font(bold=True, color="FFFFFF")
    symbols = list(corr_df.columns)

    ws.cell(row=1, column=1).value = ""
    for col_idx, sym in enumerate(symbols, start=2):
        cell = ws.cell(row=1, column=col_idx, value=sym)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, sym in enumerate(symbols, start=2):
        h = ws.cell(row=row_idx, column=1, value=sym)
        h.font  = header_font
        h.fill  = header_fill
        h.alignment = Alignment(horizontal="center")
        for col_idx, col_sym in enumerate(symbols, start=2):
            val = corr_df.loc[sym, col_sym]
            cell = ws.cell(row=row_idx, column=col_idx, value=round(float(val), 4))
            cell.number_format = "0.00"

    ws.freeze_panes = "B2"
    _autofit(ws)

    return _save(wb)


# ── Summary metrics export (all strategies, mirrors Summary tab) ──────────────

def export_summary_metrics(
    summary_df: "pd.DataFrame",
    strategies: list[dict],
) -> bytes:
    """
    Export all-strategies performance summary to .xlsx.

    Mirrors the Summary tab: 80+ computed metrics for every strategy,
    with Status and Contracts prepended from the strategies config.
    """
    wb = _make_workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Strategy Summary")

    if not summary_df.empty:
        status_map    = {s.get("name", ""): s.get("status", "")    for s in strategies}
        contracts_map = {s.get("name", ""): s.get("contracts", 1) for s in strategies}

        df = summary_df.copy().reset_index()
        # Normalise index column name
        if "strategy_name" in df.columns:
            df = df.rename(columns={"strategy_name": "Strategy"})
        elif df.columns[0] != "Strategy":
            df = df.rename(columns={df.columns[0]: "Strategy"})

        df.insert(1, "Status",    [status_map.get(n, "")    for n in df["Strategy"]])
        df.insert(2, "Contracts", [contracts_map.get(n, 1) for n in df["Strategy"]])

        # Convert date columns to strings
        for col in df.columns:
            if pd.api.types.is_datetime64_any_dtype(df[col]):
                df[col] = df[col].dt.strftime("%Y-%m-%d")
            elif df[col].dtype == object:
                try:
                    sample = df[col].dropna().iloc[0]
                    if hasattr(sample, "strftime"):
                        df[col] = df[col].apply(
                            lambda v: v.strftime("%Y-%m-%d") if v is not None and hasattr(v, "strftime") else v
                        )
                except (IndexError, TypeError):
                    pass

        _df_to_sheet(ws, df, title="All Strategies — Performance Summary")
    else:
        ws.append(["No summary metrics available."])

    ws_meta = wb.create_sheet("Export Info")
    ws_meta.append(["Portfolio Tracker v2 — Strategy Summary Export"])
    ws_meta.append(["Generated",   str(date.today())])
    ws_meta.append(["Strategies",  len(summary_df) if not summary_df.empty else 0])
    _autofit(ws_meta)

    return _save(wb)


# ── Equity curves export (individual strategies + portfolio total) ─────────────

def export_equity_curves(daily_pnl: "pd.DataFrame") -> bytes:
    """
    Export individual strategy and portfolio equity curves to .xlsx.

    Sheets:
      Individual Daily PnL   — date + each strategy's daily PnL
      Individual Cumulative  — date + each strategy's cumulative equity
      Portfolio Total        — date + total daily PnL + cumulative equity
    """
    wb = _make_workbook()
    wb.remove(wb.active)

    def _matrix(ws, df: "pd.DataFrame", title: str) -> None:
        out = df.copy()
        out.index = out.index.strftime("%Y-%m-%d")
        out = out.reset_index().rename(columns={"index": "Date"})
        _df_to_sheet(ws, out.round(2), title=title)

    _matrix(wb.create_sheet("Individual Daily PnL"),  daily_pnl,            "Individual Strategy Daily PnL ($)")
    _matrix(wb.create_sheet("Individual Cumulative"), daily_pnl.cumsum(),   "Individual Strategy Cumulative Equity ($)")

    port_pnl    = daily_pnl.sum(axis=1)
    port_equity = port_pnl.cumsum()
    port_df = pd.DataFrame({
        "Date":                   port_pnl.index.strftime("%Y-%m-%d"),
        "Daily PnL ($)":          port_pnl.values.round(2),
        "Cumulative Equity ($)":  port_equity.values.round(2),
    })
    _df_to_sheet(wb.create_sheet("Portfolio Total"), port_df, title="Portfolio Total Equity Curve")

    ws_meta = wb.create_sheet("Export Info")
    ws_meta.append(["Portfolio Tracker v2 — Equity Curves Export"])
    ws_meta.append(["Generated",     str(date.today())])
    ws_meta.append(["Strategies",    len(daily_pnl.columns)])
    ws_meta.append(["Trading Days",  len(daily_pnl)])
    _autofit(ws_meta)

    return _save(wb)


# ── Backtest period export ────────────────────────────────────────────────────

def export_backtest_period(
    port_pnl: "pd.Series",
    strat_pnl: "pd.DataFrame",
    start_date,
    end_date,
    metrics: dict,
) -> bytes:
    """
    Export a filtered backtest period to .xlsx.

    Sheets:
      Summary Metrics  — key period stats
      Daily PnL        — individual strategy + portfolio total daily PnL
      Cumulative Equity — individual + portfolio cumulative curves
      Monthly PnL      — portfolio monthly aggregation
    """
    from openpyxl.styles import Font

    wb = _make_workbook()
    wb.remove(wb.active)

    # Summary metrics
    ws_m = wb.create_sheet("Summary Metrics")
    ws_m.append(["Portfolio Tracker v2 — Backtest Period"])
    ws_m.append(["Period",    f"{start_date} → {end_date}"])
    ws_m.append(["Generated", str(date.today())])
    ws_m.append([])
    ws_m.append(["Metric", "Value"])
    ws_m.cell(row=5, column=1).font = Font(bold=True)
    ws_m.cell(row=5, column=2).font = Font(bold=True)
    for k, v in metrics.items():
        ws_m.append([k, v])
    _autofit(ws_m)

    # Daily PnL (all strategies + portfolio total)
    combined = strat_pnl.copy()
    combined["Portfolio Total"] = port_pnl
    dates_str = combined.index.strftime("%Y-%m-%d")

    daily_out = combined.copy()
    daily_out.index = dates_str
    daily_out = daily_out.reset_index().rename(columns={"index": "Date"})
    _df_to_sheet(wb.create_sheet("Daily PnL"), daily_out.round(2),
                 title=f"Daily PnL — {start_date} to {end_date}")

    cum_out = combined.cumsum().copy()
    cum_out.index = dates_str
    cum_out = cum_out.reset_index().rename(columns={"index": "Date"})
    _df_to_sheet(wb.create_sheet("Cumulative Equity"), cum_out.round(2),
                 title=f"Cumulative Equity — {start_date} to {end_date}")

    # Monthly PnL
    monthly = port_pnl.resample("ME").sum()
    monthly_df = pd.DataFrame({
        "Month":              monthly.index.strftime("%Y-%m"),
        "Portfolio PnL ($)": monthly.values.round(2),
    })
    _df_to_sheet(wb.create_sheet("Monthly PnL"), monthly_df, title="Monthly Portfolio PnL")

    return _save(wb)


# ── What-If backtest export ───────────────────────────────────────────────────

def export_whatif_backtest(
    window: "pd.DataFrame",
    total_pnl: "pd.Series",
    start_date,
    end_date,
    metrics: dict,
    contracts_override: dict,
) -> bytes:
    """
    Export a What-If backtest composition + results to .xlsx.

    Sheets:
      Composition      — strategies + contract counts
      Summary Metrics  — period stats
      Daily PnL        — individual strategy + portfolio total daily PnL
      Cumulative Equity — cumulative curves
      Monthly PnL      — monthly aggregation
    """
    from openpyxl.styles import Font

    wb = _make_workbook()
    wb.remove(wb.active)

    # Composition
    ws_comp = wb.create_sheet("Composition")
    ws_comp.append(["Strategy", "Contracts"])
    ws_comp.cell(row=1, column=1).font = Font(bold=True)
    ws_comp.cell(row=1, column=2).font = Font(bold=True)
    for name, c in contracts_override.items():
        ws_comp.append([name, c])
    _autofit(ws_comp)

    # Summary metrics
    ws_m = wb.create_sheet("Summary Metrics")
    ws_m.append(["Portfolio Tracker v2 — What-If Backtest"])
    ws_m.append(["Period",    f"{start_date} → {end_date}"])
    ws_m.append(["Generated", str(date.today())])
    ws_m.append([])
    ws_m.append(["Metric", "Value"])
    ws_m.cell(row=5, column=1).font = Font(bold=True)
    ws_m.cell(row=5, column=2).font = Font(bold=True)
    for k, v in metrics.items():
        ws_m.append([k, v])
    _autofit(ws_m)

    # Daily PnL
    combined = window.copy()
    combined["Portfolio Total"] = total_pnl
    dates_str = combined.index.strftime("%Y-%m-%d")

    daily_out = combined.copy()
    daily_out.index = dates_str
    daily_out = daily_out.reset_index().rename(columns={"index": "Date"})
    _df_to_sheet(wb.create_sheet("Daily PnL"), daily_out.round(2),
                 title=f"What-If Daily PnL — {start_date} to {end_date}")

    cum_out = combined.cumsum().copy()
    cum_out.index = dates_str
    cum_out = cum_out.reset_index().rename(columns={"index": "Date"})
    _df_to_sheet(wb.create_sheet("Cumulative Equity"), cum_out.round(2),
                 title=f"What-If Cumulative Equity — {start_date} to {end_date}")

    # Monthly PnL
    monthly = total_pnl.resample("ME").sum()
    monthly_df = pd.DataFrame({
        "Month":              monthly.index.strftime("%Y-%m"),
        "Portfolio PnL ($)": monthly.values.round(2),
    })
    _df_to_sheet(wb.create_sheet("Monthly PnL"), monthly_df, title="Monthly Portfolio PnL")

    return _save(wb)


# ── Strategy detail export ────────────────────────────────────────────────────

def export_strategy_detail(
    strategy_name: str,
    scaled_pnl: "pd.Series",
    oos_start,
    summary_row: "pd.Series | None" = None,
) -> bytes:
    """
    Export full strategy drill-down to .xlsx.

    Sheets:
      Daily PnL        — date, daily PnL, cumulative equity, IS/OOS flag
      Monthly PnL      — monthly aggregation with IS/OOS flag
      IS vs OOS Metrics — key metrics for each period
      All Metrics      — all 80+ computed metrics (if available)
    """
    import numpy as np

    wb = _make_workbook()
    wb.remove(wb.active)

    oos_ts = pd.Timestamp(oos_start) if oos_start else None

    # Daily PnL
    pnl_df = pd.DataFrame({
        "Date":                   scaled_pnl.index.strftime("%Y-%m-%d"),
        "Daily PnL ($)":          scaled_pnl.values.round(2),
        "Cumulative Equity ($)":  scaled_pnl.cumsum().values.round(2),
        "Period": [
            "OOS" if (oos_ts is not None and ts >= oos_ts) else "IS"
            for ts in scaled_pnl.index
        ],
    })
    _df_to_sheet(wb.create_sheet("Daily PnL"), pnl_df,
                 title=f"{strategy_name} — Daily PnL")

    # Monthly PnL
    monthly = scaled_pnl.resample("ME").sum()
    monthly_df = pd.DataFrame({
        "Month":    monthly.index.strftime("%Y-%m"),
        "PnL ($)":  monthly.values.round(2),
        "Period": [
            "OOS" if (oos_ts is not None and ts >= oos_ts) else "IS"
            for ts in monthly.index
        ],
    })
    _df_to_sheet(wb.create_sheet("Monthly PnL"), monthly_df,
                 title=f"{strategy_name} — Monthly PnL")

    # IS vs OOS metrics
    def _calc(pnl: "pd.Series", label: str) -> dict:
        if pnl.empty:
            return {"Period": label, "Trading Days": 0}
        eq = pnl.cumsum()
        peak = eq.cummax()
        dd = peak - eq
        n_years = max(len(pnl) / 252.0, 1e-3)
        total   = float(pnl.sum())
        ann     = total / n_years
        max_dd  = float(dd.max())
        monthly_s = pnl.resample("ME").sum()
        win_rate  = float((monthly_s > 0).mean()) if len(monthly_s) > 0 else 0.0
        std_m     = float(monthly_s.std()) if len(monthly_s) > 1 else 0.0
        sharpe    = (float(monthly_s.mean()) / std_m * np.sqrt(12)) if std_m > 1e-9 else 0.0
        rtd       = abs(total / max_dd) if max_dd > 0 else 0.0
        return {
            "Period":             label,
            "Total PnL ($)":      round(total,    2),
            "Annual PnL ($)":     round(ann,      2),
            "Max Drawdown ($)":   round(max_dd,   2),
            "R:DD":               round(rtd,      3),
            "Monthly Win Rate":   f"{win_rate:.1%}",
            "Sharpe (Monthly)":   round(sharpe,   3),
            "Trading Days":       len(pnl),
        }

    is_pnl  = scaled_pnl[scaled_pnl.index <  oos_ts] if oos_ts is not None else scaled_pnl
    oos_pnl = scaled_pnl[scaled_pnl.index >= oos_ts] if oos_ts is not None else pd.Series(dtype=float)
    metrics_df = pd.DataFrame([_calc(is_pnl, "In-Sample (IS)"), _calc(oos_pnl, "Out-of-Sample (OOS)")])
    _df_to_sheet(wb.create_sheet("IS vs OOS Metrics"), metrics_df,
                 title=f"{strategy_name} — IS vs OOS Metrics")

    # All walkforward metrics
    if summary_row is not None:
        all_df = pd.DataFrame([summary_row]).reset_index(drop=True)
        for col in all_df.columns:
            if pd.api.types.is_datetime64_any_dtype(all_df[col]):
                all_df[col] = all_df[col].dt.strftime("%Y-%m-%d")
            elif all_df[col].dtype == object:
                try:
                    sample = all_df[col].dropna().iloc[0]
                    if hasattr(sample, "strftime"):
                        all_df[col] = all_df[col].apply(
                            lambda v: v.strftime("%Y-%m-%d") if v is not None and hasattr(v, "strftime") else v
                        )
                except (IndexError, TypeError):
                    pass
        _df_to_sheet(wb.create_sheet("All Metrics"), all_df,
                     title=f"{strategy_name} — All Computed Metrics")

    ws_meta = wb.create_sheet("Export Info")
    ws_meta.append(["Portfolio Tracker v2 — Strategy Detail Export"])
    ws_meta.append(["Strategy",  strategy_name])
    ws_meta.append(["Generated", str(date.today())])
    ws_meta.append(["OOS Start", str(oos_start) if oos_start else "N/A"])
    _autofit(ws_meta)

    return _save(wb)


# ── Filename helpers ──────────────────────────────────────────────────────────

def portfolio_export_filename() -> str:
    return f"portfolio_output_{date.today()}.xlsx"

def raw_data_export_filename() -> str:
    return f"portfolio_raw_data_{date.today()}.xlsx"

def mc_export_filename(label: str = "portfolio") -> str:
    return f"mc_{label.lower().replace(' ', '_')}_{date.today()}.xlsx"

def loo_export_filename() -> str:
    return f"leave_one_out_{date.today()}.xlsx"

def correlations_export_filename(mode: str = "Normal") -> str:
    return f"correlations_{mode.lower()}_{date.today()}.xlsx"

def summary_metrics_export_filename() -> str:
    return f"strategy_summary_{date.today()}.xlsx"

def equity_curves_export_filename() -> str:
    return f"equity_curves_{date.today()}.xlsx"

def backtest_period_export_filename(start, end) -> str:
    return f"backtest_{start}_{end}_{date.today()}.xlsx"

def whatif_backtest_export_filename() -> str:
    return f"whatif_backtest_{date.today()}.xlsx"

def strategy_detail_export_filename(name: str) -> str:
    safe = name.replace(" ", "_").replace("/", "-")[:40]
    return f"strategy_detail_{safe}_{date.today()}.xlsx"
